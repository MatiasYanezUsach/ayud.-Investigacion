#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar reporte Excel de resultados del experimento
Se ejecuta automáticamente después de cada grupo experimental
"""

import os
import csv
import re
import platform
import subprocess
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_ram_info():
    """Obtiene información de RAM del sistema"""
    try:
        # Intentar usar psutil si está disponible
        import psutil
        ram_total = psutil.virtual_memory().total
        ram_total_gb = ram_total / (1024 ** 3)
        ram_available_gb = psutil.virtual_memory().available / (1024 ** 3)
        return f"{ram_total_gb:.2f} GB (Disponible: {ram_available_gb:.2f} GB)"
    except ImportError:
        # Fallback: usar comandos del sistema
        try:
            if platform.system() == "Windows":
                # Windows: usar wmic
                result = subprocess.run(
                    ['wmic', 'computersystem', 'get', 'TotalPhysicalMemory'],
                    capture_output=True,
                    text=True,
                    check=True
                )
                lines = result.stdout.strip().split('\n')
                if len(lines) > 1:
                    ram_bytes = int(lines[1].strip())
                    ram_gb = ram_bytes / (1024 ** 3)
                    return f"{ram_gb:.2f} GB"
            else:
                # Linux/Mac: usar /proc/meminfo o sysctl
                try:
                    with open('/proc/meminfo', 'r') as f:
                        for line in f:
                            if line.startswith('MemTotal:'):
                                ram_kb = int(line.split()[1])
                                ram_gb = ram_kb / (1024 ** 2)
                                return f"{ram_gb:.2f} GB"
                except:
                    pass
        except:
            pass
        return "N/A"

def parse_float(value):
    """Convierte un valor a float, manejando comas como separador decimal"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Reemplazar coma por punto para decimales
    return float(str(value).replace(',', '.'))

def parse_instance_name(instance_name):
    """Parsea el nombre de instancia para extraer información"""
    name = instance_name.replace('.txt', '')
    
    # Formato RC101_20_02: RC = tipo combinado (debe ir antes de R)
    if name.startswith('RC') and '_' in name:
        parts = name.split('_')
        if len(parts) >= 3:
            try:
                return {
                    'tipo': 'RC (Random-Clustered)',
                    'numero': parts[0][2:],  # Quitar "RC"
                    'nodos': parts[1],
                    'variante': parts[2] if len(parts) > 2 else 'N/A'
                }
            except:
                pass
    
    # Formato R161_15_80: R = tipo, 161 = clientes, 15 = nodos, 80 = capacidad
    if name.startswith('R') and '_' in name and not name.startswith('RC'):
        parts = name.split('_')
        if len(parts) >= 3:
            try:
                return {
                    'tipo': 'R (Random)',
                    'clientes': parts[0][1:] if parts[0][0] == 'R' else parts[0],
                    'nodos': parts[1],
                    'capacidad': parts[2] if len(parts) > 2 else 'N/A'
                }
            except:
                pass
    
    # Formato C101_20_02: C = tipo, 101 = número, 20 = nodos, 02 = variante
    if name.startswith('C') and '_' in name:
        parts = name.split('_')
        if len(parts) >= 3:
            try:
                return {
                    'tipo': 'C (Clustered)',
                    'numero': parts[0][1:],
                    'nodos': parts[1],
                    'variante': parts[2] if len(parts) > 2 else 'N/A'
                }
            except:
                pass
    
    # Formato RC101_20_02: RC = tipo combinado
    if name.startswith('RC') and '_' in name:
        parts = name.split('_')
        if len(parts) >= 3:
            try:
                return {
                    'tipo': 'RC (Random-Clustered)',
                    'numero': parts[0][2:],
                    'nodos': parts[1],
                    'variante': parts[2] if len(parts) > 2 else 'N/A'
                }
            except:
                pass
    
    # Formato 3C_20_66-02: 3C = tipo, 20 = nodos, 66 = capacidad, 02 = variante
    if name.startswith('3C') and '_' in name:
        parts = name.split('_')
        if len(parts) >= 3:
            try:
                nodos = parts[1]  # "20" o "40"
                cap_var = parts[2].split('-') if '-' in parts[2] else [parts[2], '']
                return {
                    'tipo': '3C',
                    'nodos': nodos,
                    'capacidad': cap_var[0] if len(cap_var) > 0 else 'N/A',
                    'variante': cap_var[1] if len(cap_var) > 1 else 'N/A'
                }
            except:
                pass
    
    # Formato Mitra-1-01, CON3-3, SCA3-5, CMT1x, CMT1y
    if '-' in name:
        parts = name.split('-')
        return {
            'tipo': parts[0],
            'param1': parts[1] if len(parts) > 1 else 'N/A',
            'param2': parts[2] if len(parts) > 2 else 'N/A'
        }
    
    return {'tipo': 'Desconocido', 'info': name}

def count_instances():
    """Cuenta las instancias en data/evolution"""
    evolution_path = "data/evolution"
    if os.path.exists(evolution_path):
        files = [f for f in os.listdir(evolution_path) if f.endswith('.txt')]
        return len(files)
    return 0

def read_baseline():
    """Lee el archivo de baseline por instancia"""
    baseline_file = "out/baseline/cplex_baseline_per_instance.csv"
    if not os.path.exists(baseline_file):
        print(f"ADVERTENCIA: No se encuentra {baseline_file}")
        return {}
    
    baseline = {}
    with open(baseline_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            instance = row.get('Instance', '').strip()
            # El CSV tiene comas como separador decimal, pero csv.DictReader las interpreta como separadores de campo
            # Ejemplo: "0,166966" se divide en '0' y '166966'
            time_str = row.get('BaselineTimeSeconds', '0').strip()
            
            # Si hay una clave None, significa que el valor se dividió
            if None in row:
                # Reconstruir: unir el valor principal con los valores adicionales
                extra_values = row[None]
                if isinstance(extra_values, list):
                    time_str = time_str + ',' + ','.join(extra_values)
                else:
                    time_str = time_str + ',' + str(extra_values)
            
            baseline[instance] = parse_float(time_str)
    
    return baseline

def parse_statistics_file(stats_file):
    """Parsea el archivo de estadísticas de CPLEX"""
    if not os.path.exists(stats_file):
        return None
    
    stats = {}
    with open(stats_file, 'r', encoding='utf-8') as f:
        content = f.read()
        
        # Extraer estadísticas globales
        indiv_match = re.search(r'Individuos evaluados:\s+(\d+)', content)
        calls_match = re.search(r'Llamadas totales a CPLEX:\s+([\d,\.]+)', content)
        time_match = re.search(r'Tiempo total usado:\s+([\d,\.]+)', content)
        pct_match = re.search(r'Porcentaje configurado:\s+([\d,\.]+)%', content)
        avg_calls_match = re.search(r'Promedio llamadas/indiv:\s+([\d,\.]+)', content)
        avg_time_match = re.search(r'Promedio tiempo/indiv:\s+([\d,\.]+)', content)
        
        stats['individuos'] = int(indiv_match.group(1)) if indiv_match else 0
        stats['llamadas'] = int(parse_float(calls_match.group(1))) if calls_match else 0
        stats['tiempo_total'] = parse_float(time_match.group(1)) if time_match else 0.0
        stats['porcentaje'] = parse_float(pct_match.group(1)) if pct_match else 0.0
        stats['prom_llamadas_indiv'] = parse_float(avg_calls_match.group(1)) if avg_calls_match else 0.0
        stats['prom_tiempo_indiv'] = parse_float(avg_time_match.group(1)) if avg_time_match else 0.0
    
    return stats

def read_best_fitness(fitness_file):
    """Lee el archivo BestFitness.csv"""
    if not os.path.exists(fitness_file):
        return []
    
    data = []
    with open(fitness_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            data.append({
                'Gen': int(row.get('Gen', 0)),
                'Standarized': parse_float(row.get('Standarized', 0)),
                'Ajusted': parse_float(row.get('Ajusted', 0)),
                'Hits': int(row.get('Hits', 0))
            })
    
    return data

def read_stats_prom_mej(stats_file):
    """Lee el archivo EstadisticaProm&Mej.csv"""
    if not os.path.exists(stats_file):
        return []
    
    data = []
    with open(stats_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        for row in reader:
            data.append({
                'Gen': int(row.get('Gen', 0)),
                'AvgSize': parse_float(row.get('AvgSize', 0)),
                'AvgERP': parse_float(row.get('AvgERP', 0)),
                'AvgFitness': parse_float(row.get('AvgFITNESS', 0)),
                'BestSize': int(row.get('BestSize', 0)) if row.get('BestSize', '').strip() else 0,
                'BestERP': parse_float(row.get('BestERP', 0)),
                'BestFitness': parse_float(row.get('BestFITNESS', 0)),
                'AvgHits': parse_float(row.get('AvgHits', 0)),
                'BestHits': int(row.get('BestHits', 0)) if row.get('BestHits', '').strip() else 0
            })
    
    return data

def find_all_executions(group_num=None):
    """Encuentra todas las ejecuciones en out/results/"""
    
    # Si se especifica un grupo, buscar en out/results/grupo{N}/
    if group_num is not None:
        results_dir = f"out/results/grupo{group_num}"
    else:
        results_dir = "out/results"
    
    if not os.path.exists(results_dir):
        return []
    
    executions = []
    for item in os.listdir(results_dir):
        if item.startswith('evolution'):
            try:
                exec_num = int(item.replace('evolution', ''))
                exec_dir = os.path.join(results_dir, item)
                if os.path.isdir(exec_dir):
                    executions.append(exec_num)
            except ValueError:
                continue
    
    return sorted(executions)

def generate_excel(group_num=None):
    """Genera el archivo Excel con todos los resultados"""
    
    # Determinar nombre del archivo
    if group_num is not None:
        excel_file = f"RESULTADOS_EXPERIMENTO_GRUPO{group_num}.xlsx"
    else:
        excel_file = "RESULTADOS_EXPERIMENTO.xlsx"
    
    print(f"\nGenerando reporte Excel: {excel_file}")
    print("=" * 80)
    
    # Leer baseline
    print("Leyendo baseline...")
    baseline = read_baseline()
    
    # Contar instancias
    num_instances = count_instances()
    
    # Encontrar todas las ejecuciones
    print("Buscando ejecuciones...")
    executions = find_all_executions(group_num)  # Pasar group_num aqui
    
    if not executions:
        if group_num is not None:
            print(f"ERROR: No se encontraron ejecuciones en out/results/grupo{group_num}/")
        else:
            print("ERROR: No se encontraron ejecuciones en out/results/")
        return False
    
    print(f"Encontradas {len(executions)} ejecuciones: {executions}")
    
    # Determinar el directorio base de resultados
    if group_num is not None:
        results_base_dir = f"out/results/grupo{group_num}"
    else:
        results_base_dir = "out/results"
    
    # Crear workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remover hoja por defecto
    
    # ===== HOJA 0: Configuración del Experimento =====
    print("Creando hoja: Configuración del Experimento...")
    ws_config = wb.create_sheet("Configuración", 0)  # Primera hoja
    
    row = 1
    ws_config.merge_cells(f'A{row}:C{row}')
    cell = ws_config.cell(row, 1)
    cell.value = f"CONFIGURACIÓN DEL EXPERIMENTO - GRUPO {group_num if group_num is not None else 'N/A'}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    
    # Información de la máquina
    ws_config.cell(row, 1).value = "INFORMACIÓN DE LA MÁQUINA"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "Sistema Operativo:"
    ws_config.cell(row, 2).value = platform.system() + " " + platform.release()
    row += 1
    ws_config.cell(row, 1).value = "Arquitectura:"
    ws_config.cell(row, 2).value = platform.machine()
    row += 1
    ws_config.cell(row, 1).value = "Procesador:"
    ws_config.cell(row, 2).value = platform.processor() if platform.processor() else "N/A"
    row += 1
    ws_config.cell(row, 1).value = "Memoria RAM:"
    ws_config.cell(row, 2).value = get_ram_info()
    row += 2
    
    # Parámetros del algoritmo evolutivo
    ws_config.cell(row, 1).value = "PARÁMETROS DEL ALGORITMO EVOLUTIVO"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "Tamaño de Población:"
    ws_config.cell(row, 2).value = "15 individuos"
    row += 1
    ws_config.cell(row, 1).value = "Generaciones:"
    ws_config.cell(row, 2).value = "100 generaciones"
    row += 1
    ws_config.cell(row, 1).value = "Elitismo:"
    ws_config.cell(row, 2).value = "2 individuos"
    row += 1
    ws_config.cell(row, 1).value = "Probabilidad Crossover:"
    ws_config.cell(row, 2).value = "70%"
    row += 1
    ws_config.cell(row, 1).value = "Probabilidad Mutación (Subtree):"
    ws_config.cell(row, 2).value = "15%"
    row += 1
    ws_config.cell(row, 1).value = "Probabilidad Mutación (One Node):"
    ws_config.cell(row, 2).value = "15%"
    row += 1
    ws_config.cell(row, 1).value = "Selección:"
    ws_config.cell(row, 2).value = "Tournament Selection (tamaño 4)"
    row += 2
    
    # Instancias
    ws_config.cell(row, 1).value = "INSTANCIAS DE EVALUACIÓN"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "Número de Instancias:"
    ws_config.cell(row, 2).value = f"{num_instances} instancias"
    ws_config.cell(row, 3).value = "(Cada individuo se evalúa con todas las instancias)"
    row += 2
    
    # Fitness
    ws_config.cell(row, 1).value = "FITNESS"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "Fitness Estandarizado:"
    ws_config.cell(row, 2).value = "Error Relativo Promedio (ERP)"
    row += 1
    ws_config.cell(row, 1).value = "Fitness Ajustado:"
    ws_config.cell(row, 2).value = "1 / (1 + Fitness Estandarizado)"
    row += 1
    ws_config.cell(row, 1).value = "Fitness Usado:"
    ws_config.cell(row, 2).value = "Fitness Estandarizado (ERP)"
    ws_config.cell(row, 2).font = Font(bold=True, color="FF0000")
    row += 1
    ws_config.cell(row, 1).value = "ERP (Error Relativo Promedio):"
    ws_config.cell(row, 2).value = "Promedio del error relativo de todas las instancias"
    row += 1
    ws_config.cell(row, 1).value = "Error Relativo por Instancia:"
    ws_config.cell(row, 2).value = "|óptimo - solución_obtenida| / óptimo"
    row += 2
    
    # Notas importantes
    ws_config.cell(row, 1).value = "NOTAS IMPORTANTES"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "• Los 'Individuos' en 'Estadísticas por Ejecución' son el total de individuos evaluados durante toda la ejecución (no el tamaño de población)"
    ws_config.merge_cells(f'A{row}:C{row}')
    row += 1
    ws_config.cell(row, 1).value = "• El 'Best Fitness' es del MEJOR individuo de cada generación"
    ws_config.merge_cells(f'A{row}:C{row}')
    row += 1
    ws_config.cell(row, 1).value = "• 'AvgERP' es el Error Relativo Promedio de TODA la población en esa generación"
    ws_config.merge_cells(f'A{row}:C{row}')
    row += 1
    ws_config.cell(row, 1).value = "• 'BestERP' es el Error Relativo Promedio del MEJOR individuo de esa generación"
    ws_config.merge_cells(f'A{row}:C{row}')
    row += 1
    ws_config.cell(row, 1).value = f"• Cada individuo se evalúa con {num_instances} instancias de prueba"
    ws_config.merge_cells(f'A{row}:C{row}')
    row += 1
    ws_config.cell(row, 1).value = "• El formato de números usa punto (.) como separador decimal (formato inglés)"
    ws_config.merge_cells(f'A{row}:C{row}')
    
    # Ajustar ancho de columnas
    ws_config.column_dimensions['A'].width = 35
    ws_config.column_dimensions['B'].width = 40
    ws_config.column_dimensions['C'].width = 50
    
    # ===== HOJA 1: Resumen General =====
    print("Creando hoja: Resumen General...")
    ws_summary = wb.create_sheet("Resumen General")
    
    row = 1
    ws_summary.merge_cells(f'A{row}:E{row}')
    cell = ws_summary.cell(row, 1)
    cell.value = f"RESUMEN GENERAL - EXPERIMENTO GRUPO {group_num if group_num is not None else 'N/A'}"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    
    # Baseline CPLEX
    ws_summary.cell(row, 1).value = "BASELINE CPLEX (Tiempo por Instancia - CPLEX Puro)"
    ws_summary.cell(row, 1).font = Font(bold=True)
    row += 1
    
    if baseline:
        times = list(baseline.values())
        ws_summary.cell(row, 1).value = "Promedio:"
        ws_summary.cell(row, 2).value = f"{sum(times)/len(times):.3f} segundos"
        row += 1
        ws_summary.cell(row, 1).value = "Mínimo:"
        ws_summary.cell(row, 2).value = f"{min(times):.3f} segundos"
        row += 1
        ws_summary.cell(row, 1).value = "Máximo:"
        ws_summary.cell(row, 2).value = f"{max(times):.3f} segundos"
        row += 1
        ws_summary.cell(row, 1).value = "Desviación Estándar:"
        if len(times) > 1:
            mean = sum(times) / len(times)
            variance = sum((t - mean) ** 2 for t in times) / (len(times) - 1)
            std_dev = variance ** 0.5
            ws_summary.cell(row, 2).value = f"{std_dev:.3f} segundos"
        row += 2
    
    # Estadísticas consolidadas
    ws_summary.cell(row, 1).value = "ESTADISTICAS CONSOLIDADAS (Todas las Ejecuciones)"
    ws_summary.cell(row, 1).font = Font(bold=True)
    row += 1
    
    # Recopilar estadísticas de todas las ejecuciones
    all_stats = []
    for exec_num in executions:
        stats_file = f"{results_base_dir}/evolution{exec_num}/job.{exec_num}.CplexUsage.statistics.txt"
        stats = parse_statistics_file(stats_file)
        if stats:
            all_stats.append(stats)
    
    # Determinar el presupuesto esperado basándose en el número de grupo
    expected_budget_pct = None
    if group_num is not None:
        budget_map = {0: 0.0, 1: 10.0, 2: 25.0, 3: 50.0, 4: 75.0, 5: 100.0}
        expected_budget_pct = budget_map.get(group_num)
    
    # Calcular individuos evaluados desde EstadisticaProm&Mej.csv (SIEMPRE disponible, incluso para Grupo 0)
    total_indiv_from_csv = 0
    all_best_erp_final = []
    all_avg_erp_final = []
    all_best_fitness_final = []
    all_avg_fitness_final = []
    
    for exec_num in executions:
        stats_file_csv = f"{results_base_dir}/evolution{exec_num}/job.{exec_num}.EstadisticaProm&Mej.csv"
        try:
            with open(stats_file_csv, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')  # CSV usa punto y coma como delimitador
                rows = list(reader)
                if rows:
                    # Última fila = generación final
                    last_row = rows[-1]
                    # IndivEvals está en la columna 'Evaluated' o puede estar calculado como Gen * pop_size
                    # Para este experimento: 100 generaciones × 15 individuos = 1500 por ejecución
                    gen = parse_float(last_row.get('Gen', 0))
                    if gen > 0:
                        total_indiv_from_csv += gen * 15  # 15 = tamaño de población
                    
                    # Capturar ERP y Fitness finales
                    all_best_erp_final.append(parse_float(last_row.get('BestERP', 0)))
                    all_avg_erp_final.append(parse_float(last_row.get('AvgERP', 0)))
                    all_best_fitness_final.append(parse_float(last_row.get('BestFITNESS', 0)))
                    all_avg_fitness_final.append(parse_float(last_row.get('AvgFITNESS', 0)))
        except Exception as e:
            print(f"⚠️  Advertencia: No se pudo leer {stats_file_csv}: {e}")
    
    if all_stats:
        total_indiv = sum(s['individuos'] for s in all_stats)
        total_calls = sum(s['llamadas'] for s in all_stats)
        total_time = sum(s['tiempo_total'] for s in all_stats)
        avg_pct = sum(s['porcentaje'] for s in all_stats) / len(all_stats) if all_stats else 0
        
        ws_summary.cell(row, 1).value = "Total Ejecuciones:"
        ws_summary.cell(row, 2).value = len(all_stats)
        row += 1
        ws_summary.cell(row, 1).value = "Total Individuos Evaluados:"
        ws_summary.cell(row, 2).value = total_indiv
        ws_summary.cell(row, 3).value = "(Suma de todos los individuos evaluados en todas las ejecuciones)"
        row += 1
        ws_summary.cell(row, 1).value = "Total Llamadas CPLEX:"
        ws_summary.cell(row, 2).value = total_calls
        row += 1
        ws_summary.cell(row, 1).value = "Tiempo Total CPLEX:"
        ws_summary.cell(row, 2).value = f"{total_time:.2f} segundos"
        row += 1
        ws_summary.cell(row, 1).value = "Presupuesto CPLEX:"
        
        # Usar el presupuesto esperado si se especificó un grupo, sino usar el calculado
        if expected_budget_pct is not None:
            ws_summary.cell(row, 2).value = f"{expected_budget_pct:.1f}% del baseline por instancia"
            # Advertir si hay discrepancia
            if abs(avg_pct - expected_budget_pct) > 1.0:
                ws_summary.cell(row, 3).value = f"ADVERTENCIA: Los archivos muestran {avg_pct:.1f}% - Verifique que las ejecuciones correspondan al grupo correcto"
                ws_summary.cell(row, 3).font = Font(italic=True, color="FF0000")
        else:
            ws_summary.cell(row, 2).value = f"{avg_pct:.1f}% del baseline por instancia"
        row += 1
        
        # Agregar estadísticas de rendimiento evolutivo
        if all_best_erp_final:
            row += 1
            ws_summary.cell(row, 1).value = "RENDIMIENTO EVOLUTIVO (Generación Final)"
            ws_summary.cell(row, 1).font = Font(bold=True)
            row += 1
            
            ws_summary.cell(row, 1).value = "Mejor ERP Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_best_erp_final) / len(all_best_erp_final):.6f}"
            ws_summary.cell(row, 3).value = "(Promedio del mejor ERP de cada ejecución)"
            row += 1
            
            ws_summary.cell(row, 1).value = "ERP Poblacional Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_avg_erp_final) / len(all_avg_erp_final):.6f}"
            ws_summary.cell(row, 3).value = "(Promedio del ERP de la población en cada ejecución)"
            row += 1
            
            ws_summary.cell(row, 1).value = "Mejor Fitness Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_best_fitness_final) / len(all_best_fitness_final):.6f}"
            row += 1
            
            ws_summary.cell(row, 1).value = "Fitness Poblacional Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_avg_fitness_final) / len(all_avg_fitness_final):.6f}"
            
    elif group_num == 0:
        # Grupo 0: Sin CPLEX - Pero SÍ tiene datos de rendimiento evolutivo
        ws_summary.cell(row, 1).value = "Total Ejecuciones:"
        ws_summary.cell(row, 2).value = len(executions)
        row += 1
        
        # Agregar individuos evaluados desde CSV
        if total_indiv_from_csv > 0:
            ws_summary.cell(row, 1).value = "Total Individuos Evaluados:"
            ws_summary.cell(row, 2).value = int(total_indiv_from_csv)
            ws_summary.cell(row, 3).value = "(Suma de todos los individuos evaluados en todas las ejecuciones)"
            row += 1
        
        ws_summary.cell(row, 1).value = "Total Llamadas CPLEX:"
        ws_summary.cell(row, 2).value = 0
        ws_summary.cell(row, 3).value = "(Grupo de control - No usa CPLEX)"
        row += 1
        
        ws_summary.cell(row, 1).value = "Tiempo Total CPLEX:"
        ws_summary.cell(row, 2).value = "0.00 segundos"
        ws_summary.cell(row, 3).value = "(Grupo de control - No usa CPLEX)"
        row += 1
        
        ws_summary.cell(row, 1).value = "Presupuesto CPLEX:"
        ws_summary.cell(row, 2).value = "0.0% del baseline por instancia (Sin CPLEX)"
        ws_summary.cell(row, 3).value = "Grupo de control - Algoritmo Genético Puro"
        row += 1
        
        # Agregar estadísticas de rendimiento evolutivo para Grupo 0
        if all_best_erp_final:
            row += 1
            ws_summary.cell(row, 1).value = "RENDIMIENTO EVOLUTIVO (Generación Final)"
            ws_summary.cell(row, 1).font = Font(bold=True)
            ws_summary.cell(row, 3).value = "IMPORTANTE: Estas métricas sirven como BASELINE para comparar contra grupos con CPLEX"
            ws_summary.cell(row, 3).font = Font(italic=True, color="0000FF")
            row += 1
            
            ws_summary.cell(row, 1).value = "Mejor ERP Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_best_erp_final) / len(all_best_erp_final):.6f}"
            ws_summary.cell(row, 3).value = "(Promedio del mejor ERP de cada ejecución)"
            row += 1
            
            ws_summary.cell(row, 1).value = "ERP Poblacional Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_avg_erp_final) / len(all_avg_erp_final):.6f}"
            ws_summary.cell(row, 3).value = "(Promedio del ERP de la población en cada ejecución)"
            row += 1
            
            ws_summary.cell(row, 1).value = "Mejor Fitness Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_best_fitness_final) / len(all_best_fitness_final):.6f}"
            row += 1
            
            ws_summary.cell(row, 1).value = "Fitness Poblacional Promedio:"
            ws_summary.cell(row, 2).value = f"{sum(all_avg_fitness_final) / len(all_avg_fitness_final):.6f}"
    
    # Ajustar ancho de columnas
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 50
    
    # ===== HOJA 2: Baseline por Instancia =====
    print("Creando hoja: Baseline por Instancia...")
    ws_baseline = wb.create_sheet("Baseline por Instancia")
    
    # Nota sobre la máquina
    ws_baseline.cell(1, 1).value = "NOTA: Estos tiempos fueron obtenidos resolviendo cada instancia con CPLEX puro (sin programación genética). Ver información de la máquina en la hoja 'Configuración'."
    ws_baseline.merge_cells(f'A1:E1')
    ws_baseline.cell(1, 1).font = Font(italic=True, color="0000FF")
    ws_baseline.cell(1, 1).alignment = Alignment(wrap_text=True)
    
    headers = ["Instancia", "Tiempo Baseline (s)", "Tipo", "Clientes/Nodos", "Capacidad/Variante"]
    for col, header in enumerate(headers, 1):
        cell = ws_baseline.cell(2, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Ordenar por tiempo (de menor a mayor)
    sorted_baseline = sorted(baseline.items(), key=lambda x: x[1])
    
    row = 3
    for instance, time in sorted_baseline:
        ws_baseline.cell(row, 1).value = instance
        cell_time = ws_baseline.cell(row, 2)
        cell_time.value = float(time)  # Asegurar que sea float
        cell_time.number_format = '0.000000'  # Formato con 6 decimales
        
        # Parsear nombre de instancia
        info = parse_instance_name(instance)
        ws_baseline.cell(row, 3).value = info.get('tipo', 'N/A')
        
        # Información adicional según el tipo
        if 'clientes' in info:
            ws_baseline.cell(row, 4).value = f"Clientes: {info['clientes']}, Nodos: {info['nodos']}"
            ws_baseline.cell(row, 5).value = f"Capacidad: {info.get('capacidad', 'N/A')}"
        elif 'numero' in info:
            ws_baseline.cell(row, 4).value = f"Número: {info['numero']}, Nodos: {info['nodos']}"
            ws_baseline.cell(row, 5).value = f"Variante: {info.get('variante', 'N/A')}"
        elif 'nodos' in info:
            ws_baseline.cell(row, 4).value = f"Nodos: {info['nodos']}"
            ws_baseline.cell(row, 5).value = f"Capacidad: {info.get('capacidad', 'N/A')}, Variante: {info.get('variante', 'N/A')}"
        elif 'param1' in info:
            ws_baseline.cell(row, 4).value = f"{info.get('param1', 'N/A')}"
            ws_baseline.cell(row, 5).value = f"{info.get('param2', 'N/A')}"
        else:
            ws_baseline.cell(row, 4).value = info.get('info', instance)
            ws_baseline.cell(row, 5).value = "N/A"
        
        row += 1
    
    for col in range(1, len(headers) + 1):
        ws_baseline.column_dimensions[get_column_letter(col)].width = 25
    
    # ===== HOJA 3: Estadísticas por Ejecución =====
    print("Creando hoja: Estadísticas por Ejecución...")
    ws_stats = wb.create_sheet("Estadísticas por Ejecución")
    
    # Nota explicativa
    ws_stats.cell(1, 1).value = "NOTA IMPORTANTE: 'Individuos Evaluados' = Total acumulado de individuos evaluados durante TODA la ejecución (100 generaciones). NO es el tamaño de población (que es 15). 'Llamadas CPLEX' = Total de llamadas a CPLEX durante toda la ejecución. 'Tiempo Total' = Tiempo total en segundos usado por CPLEX en toda la ejecución."
    ws_stats.merge_cells(f'A1:F1')
    ws_stats.cell(1, 1).font = Font(italic=True, color="0000FF")
    ws_stats.cell(1, 1).alignment = Alignment(wrap_text=True)
    
    headers = ["Ejecución", "Individuos Evaluados", "Llamadas CPLEX (Total)", "Tiempo Total CPLEX (s)", 
               "Prom. Llamadas/Indiv", "Prom. Tiempo/Indiv (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws_stats.cell(2, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 3
    for exec_num in executions:
        stats_file = f"{results_base_dir}/evolution{exec_num}/job.{exec_num}.CplexUsage.statistics.txt"
        stats = parse_statistics_file(stats_file)
        if stats:
            ws_stats.cell(row, 1).value = exec_num
            ws_stats.cell(row, 2).value = stats['individuos']
            ws_stats.cell(row, 3).value = stats['llamadas']
            ws_stats.cell(row, 4).value = stats['tiempo_total']
            ws_stats.cell(row, 5).value = stats['prom_llamadas_indiv']
            ws_stats.cell(row, 6).value = stats['prom_tiempo_indiv']
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_stats.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== HOJA 4: Best Fitness por Ejecución =====
    print("Creando hoja: Best Fitness por Ejecución...")
    ws_fitness = wb.create_sheet("Best Fitness por Ejecución")
    
    # Nota explicativa
    ws_fitness.cell(1, 1).value = "NOTA IMPORTANTE: Estos son los valores del MEJOR individuo de cada generación. Fitness Estandarizado = ERP (Error Relativo Promedio sobre todas las instancias). Fitness Ajustado = 1/(1+ERP). El fitness que se usa para la selección es el ESTANDARIZADO (ERP). Cada individuo se evalúa con todas las instancias de prueba."
    ws_fitness.merge_cells(f'A1:E1')
    ws_fitness.cell(1, 1).font = Font(italic=True, color="0000FF")
    ws_fitness.cell(1, 1).alignment = Alignment(wrap_text=True)
    
    headers = ["Ejecución", "Generación", "Fitness Estandarizado (ERP)", "Fitness Ajustado", "Hits"]
    for col, header in enumerate(headers, 1):
        cell = ws_fitness.cell(2, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 3
    for exec_num in executions:
        fitness_file = f"{results_base_dir}/evolution{exec_num}/job.{exec_num}.BestFitness.csv"
        fitness_data = read_best_fitness(fitness_file)
        for entry in fitness_data:
            ws_fitness.cell(row, 1).value = exec_num
            ws_fitness.cell(row, 2).value = entry['Gen']
            ws_fitness.cell(row, 3).value = entry['Standarized']  # Este es el que se usa
            ws_fitness.cell(row, 4).value = entry['Ajusted']
            ws_fitness.cell(row, 5).value = entry['Hits']  # Número de instancias donde encontró el óptimo
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_fitness.column_dimensions[get_column_letter(col)].width = 25
    
    # ===== HOJA 5: Estadísticas Promedio y Mejor =====
    print("Creando hoja: Estadísticas Promedio y Mejor...")
    ws_stats_pm = wb.create_sheet("Estadísticas Promedio y Mejor")
    
    # Nota explicativa
    ws_stats_pm.cell(1, 1).value = f"NOTA IMPORTANTE: AvgERP = Error Relativo Promedio (ERP) de TODA la población en esa generación. BestERP = Error Relativo Promedio (ERP) del MEJOR individuo de esa generación. Cada individuo se evalúa con {num_instances} instancias de prueba. El ERP se calcula como: promedio de |óptimo - solución| / óptimo sobre todas las instancias."
    ws_stats_pm.merge_cells(f'A1:J1')
    ws_stats_pm.cell(1, 1).font = Font(italic=True, color="0000FF")
    ws_stats_pm.cell(1, 1).alignment = Alignment(wrap_text=True)
    
    headers = ["Ejecución", "Gen", "AvgSize", "AvgERP", "AvgFitness", "BestSize",
               "BestERP", "BestFitness", "AvgHits", "BestHits"]
    for col, header in enumerate(headers, 1):
        cell = ws_stats_pm.cell(2, col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    row = 3
    for exec_num in executions:
        stats_file = f"{results_base_dir}/evolution{exec_num}/job.{exec_num}.EstadisticaProm&Mej.csv"
        stats_data = read_stats_prom_mej(stats_file)
        for entry in stats_data:
            ws_stats_pm.cell(row, 1).value = exec_num
            ws_stats_pm.cell(row, 2).value = entry['Gen']
            ws_stats_pm.cell(row, 3).value = entry['AvgSize']
            ws_stats_pm.cell(row, 4).value = entry['AvgERP']
            ws_stats_pm.cell(row, 5).value = entry['AvgFitness']
            ws_stats_pm.cell(row, 6).value = entry['BestSize']
            ws_stats_pm.cell(row, 7).value = entry['BestERP']
            ws_stats_pm.cell(row, 8).value = entry['BestFitness']
            ws_stats_pm.cell(row, 9).value = entry['AvgHits']
            ws_stats_pm.cell(row, 10).value = entry['BestHits']
            row += 1
    
    for col in range(1, len(headers) + 1):
        ws_stats_pm.column_dimensions[get_column_letter(col)].width = 15
    
    # Guardar archivo
    print(f"\nGuardando archivo: {excel_file}")
    wb.save(excel_file)
    print(f"[OK] Archivo Excel creado: {excel_file}")
    
    return True

if __name__ == "__main__":
    import sys
    
    group_num = None
    if len(sys.argv) > 1:
        try:
            group_num = int(sys.argv[1])
        except ValueError:
            pass
    
    success = generate_excel(group_num)
    if not success:
        exit(1)
