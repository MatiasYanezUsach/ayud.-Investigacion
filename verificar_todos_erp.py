import openpyxl

for g in [0, 1, 2, 3, 4, 5]:
    wb = openpyxl.load_workbook(f'RESULTADOS_EXPERIMENTO_GRUPO{g}.xlsx', data_only=True)
    ws = wb['Resumen General']
    
    print(f'\n=== GRUPO {g} ===')
    for i in range(10, 25):
        if ws.cell(i, 1).value and 'ERP' in str(ws.cell(i, 1).value):
            print(f'{ws.cell(i, 1).value}: {ws.cell(i, 2).value}')
    
    wb.close()
