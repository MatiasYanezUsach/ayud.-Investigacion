"""
Script para generar gráficos comparativos superpuestos entre todos los grupos experimentales.
Cada gráfico muestra las 6 curvas (una por grupo) con diferentes colores y leyendas.
"""

import openpyxl
import matplotlib.pyplot as plt
import os
import numpy as np

# Configuración de grupos y colores
GRUPOS = [0, 1, 2, 3, 4, 5]
PRESUPUESTOS = {
    0: "0% (Sin CPLEX)",
    1: "10%",
    2: "25%",
    3: "50%",
    4: "75%",
    5: "100%"
}

# Colores distintivos para cada grupo
COLORES = ['#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00', '#f781bf']

# Directorio de salida
OUTPUT_DIR = "graficos_comparativos"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Configuración de matplotlib
plt.rcParams['figure.figsize'] = (12, 8)
plt.rcParams['font.size'] = 10
plt.rcParams['axes.grid'] = True
plt.rcParams['grid.alpha'] = 0.3

def cargar_datos_grupo(grupo_num):
    """Carga los datos de un grupo desde su Excel."""
    archivo_excel = f"RESULTADOS_EXPERIMENTO_GRUPO{grupo_num}.xlsx"
    
    if not os.path.exists(archivo_excel):
        print(f"⚠️  Advertencia: No se encontró {archivo_excel}")
        return None
    
    wb = openpyxl.load_workbook(archivo_excel, data_only=True)
    
    datos = {
        'grupo': grupo_num,
        'presupuesto': PRESUPUESTOS[grupo_num],
        'baseline': {},
        'estadisticas_prom_mej': {},  # Ahora es un diccionario por generación
        'best_fitness': [],
        'estadisticas_cplex': []  # Datos de CPLEX por ejecución
    }
    
    # Leer Baseline por Instancia (datos empiezan en fila 3)
    try:
        ws_baseline = wb['Baseline por Instancia']
        for row in range(3, ws_baseline.max_row + 1):
            instancia = ws_baseline.cell(row, 1).value
            tiempo = ws_baseline.cell(row, 2).value
            if instancia and tiempo is not None:
                try:
                    datos['baseline'][str(instancia)] = float(tiempo)
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"Error leyendo baseline Grupo {grupo_num}: {e}")
    
    # Leer Estadísticas Promedio y Mejor (datos empiezan en fila 3)
    # Columnas: 1=Ejecución, 2=Gen, 3=AvgSize, 4=AvgERP, 5=AvgFitness, 6=BestSize, 7=BestERP, 8=BestFitness, 9=AvgHits, 10=BestHits
    try:
        ws_stats = wb['Estadísticas Promedio y Mejor']
        for row in range(3, ws_stats.max_row + 1):
            ejecucion = ws_stats.cell(row, 1).value
            generacion = ws_stats.cell(row, 2).value
            avg_size = ws_stats.cell(row, 3).value
            avg_erp = ws_stats.cell(row, 4).value
            avg_fitness = ws_stats.cell(row, 5).value
            best_size = ws_stats.cell(row, 6).value
            best_erp = ws_stats.cell(row, 7).value
            best_fitness = ws_stats.cell(row, 8).value
            avg_hits = ws_stats.cell(row, 9).value
            best_hits = ws_stats.cell(row, 10).value
            
            if generacion is not None:
                try:
                    gen = int(generacion)
                    if gen not in datos['estadisticas_prom_mej']:
                        datos['estadisticas_prom_mej'][gen] = {
                            'generacion': gen,
                            'avg_size_values': [],
                            'avg_erp_values': [],
                            'avg_fitness_values': [],
                            'best_size_values': [],
                            'best_erp_values': [],
                            'best_fitness_values': [],
                            'avg_hits_values': [],
                            'best_hits_values': []
                        }
                    
                    # Agregar valores (pueden ser múltiples ejecuciones por generación)
                    if avg_size is not None:
                        datos['estadisticas_prom_mej'][gen]['avg_size_values'].append(float(avg_size))
                    if avg_erp is not None:
                        datos['estadisticas_prom_mej'][gen]['avg_erp_values'].append(float(avg_erp))
                    if avg_fitness is not None:
                        datos['estadisticas_prom_mej'][gen]['avg_fitness_values'].append(float(avg_fitness))
                    if best_size is not None:
                        datos['estadisticas_prom_mej'][gen]['best_size_values'].append(float(best_size))
                    if best_erp is not None:
                        datos['estadisticas_prom_mej'][gen]['best_erp_values'].append(float(best_erp))
                    if best_fitness is not None:
                        datos['estadisticas_prom_mej'][gen]['best_fitness_values'].append(float(best_fitness))
                    if avg_hits is not None:
                        datos['estadisticas_prom_mej'][gen]['avg_hits_values'].append(float(avg_hits))
                    if best_hits is not None:
                        datos['estadisticas_prom_mej'][gen]['best_hits_values'].append(float(best_hits))
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"Error leyendo estadísticas Grupo {grupo_num}: {e}")
    
    # Calcular promedios por generación
    estadisticas_list = []
    for gen in sorted(datos['estadisticas_prom_mej'].keys()):
        stats = datos['estadisticas_prom_mej'][gen]
        estadisticas_list.append({
            'generacion': gen,
            'avg_size': np.mean(stats['avg_size_values']) if stats['avg_size_values'] else 0,
            'avg_erp': np.mean(stats['avg_erp_values']) if stats['avg_erp_values'] else 0,
            'avg_fitness': np.mean(stats['avg_fitness_values']) if stats['avg_fitness_values'] else 0,
            'best_size': np.mean(stats['best_size_values']) if stats['best_size_values'] else 0,
            'best_erp': np.mean(stats['best_erp_values']) if stats['best_erp_values'] else 0,
            'best_fitness': np.mean(stats['best_fitness_values']) if stats['best_fitness_values'] else 0,
            'avg_hits': np.mean(stats['avg_hits_values']) if stats['avg_hits_values'] else 0,
            'best_hits': np.mean(stats['best_hits_values']) if stats['best_hits_values'] else 0
        })
    datos['estadisticas_prom_mej'] = estadisticas_list
    
    # Leer Best Fitness por Ejecución (datos empiezan en fila 3)
    # Columnas: 1=Ejecución, 2=Generación, 3=Fitness Estandarizado (ERP), 4=Fitness Ajustado, 5=Hits
    try:
        ws_best = wb['Best Fitness por Ejecución']
        for row in range(3, ws_best.max_row + 1):
            ejecucion = ws_best.cell(row, 1).value
            generacion = ws_best.cell(row, 2).value
            erp = ws_best.cell(row, 3).value
            fitness_ajustado = ws_best.cell(row, 4).value
            hits = ws_best.cell(row, 5).value
            
            if generacion is not None and erp is not None:
                try:
                    datos['best_fitness'].append({
                        'ejecucion': int(ejecucion) if ejecucion is not None else 0,
                        'generacion': int(generacion),
                        'erp': float(erp),
                        'fitness_ajustado': float(fitness_ajustado) if fitness_ajustado is not None else 0,
                        'hits': float(hits) if hits is not None else 0
                    })
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"Error leyendo best fitness Grupo {grupo_num}: {e}")
    
    # Leer Estadísticas por Ejecución (para datos de CPLEX)
    # Columnas: 1=Ejecución, 2=Individuos Evaluados, 3=Llamadas CPLEX (Total), 4=Tiempo Total CPLEX (s), 5=Prom. Llamadas/Indiv, 6=Prom. Tiempo/Indiv (s)
    try:
        ws_cplex = wb['Estadísticas por Ejecución']
        for row in range(3, ws_cplex.max_row + 1):
            ejecucion = ws_cplex.cell(row, 1).value
            individuos_eval = ws_cplex.cell(row, 2).value
            llamadas_cplex = ws_cplex.cell(row, 3).value
            tiempo_cplex = ws_cplex.cell(row, 4).value
            prom_llamadas_indiv = ws_cplex.cell(row, 5).value
            prom_tiempo_indiv = ws_cplex.cell(row, 6).value
            
            if ejecucion is not None and individuos_eval is not None:
                try:
                    datos['estadisticas_cplex'].append({
                        'ejecucion': int(ejecucion),
                        'individuos_evaluados': int(individuos_eval),
                        'llamadas_cplex': float(llamadas_cplex) if llamadas_cplex is not None else 0,
                        'tiempo_cplex': float(tiempo_cplex) if tiempo_cplex is not None else 0,
                        'prom_llamadas_indiv': float(prom_llamadas_indiv) if prom_llamadas_indiv is not None else 0,
                        'prom_tiempo_indiv': float(prom_tiempo_indiv) if prom_tiempo_indiv is not None else 0
                    })
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"Error leyendo estadísticas CPLEX Grupo {grupo_num}: {e}")
    
    wb.close()
    return datos

def grafico_1_baseline_tiempos(todos_datos):
    """Gráfico 1: Tiempos baseline por instancia (comparativo)"""
    plt.figure(figsize=(14, 8))
    
    # Obtener todas las instancias únicas
    todas_instancias = set()
    for datos in todos_datos:
        if datos and 'baseline' in datos:
            todas_instancias.update(datos['baseline'].keys())
    
    instancias_ordenadas = sorted(todas_instancias)
    x_pos = np.arange(len(instancias_ordenadas))
    
    # Configurar ancho de barras
    ancho = 0.14
    offset_base = -2.5 * ancho
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['baseline']:
            continue
        
        tiempos = [datos['baseline'].get(inst, 0) for inst in instancias_ordenadas]
        offset = offset_base + idx * ancho
        
        plt.bar(x_pos + offset, tiempos, ancho, 
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Instancia', fontsize=12, fontweight='bold')
    plt.ylabel('Tiempo (segundos)', fontsize=12, fontweight='bold')
    plt.title('Comparación de Tiempos Baseline CPLEX por Instancia\n(Todos los Grupos)', 
              fontsize=14, fontweight='bold')
    plt.xticks(x_pos, instancias_ordenadas, rotation=45, ha='right')
    plt.legend(loc='best', framealpha=0.9)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '01_baseline_tiempos_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 01_baseline_tiempos_comparativo.png")

def grafico_2_llamadas_cplex(todos_datos):
    """Gráfico 2: Total de llamadas a CPLEX acumuladas por ejecución"""
    plt.figure(figsize=(12, 8))
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_cplex'] or datos['grupo'] == 0:
            continue  # Grupo 0 no usa CPLEX
        
        ejecuciones = [d['ejecucion'] for d in datos['estadisticas_cplex']]
        llamadas = [d['llamadas_cplex'] for d in datos['estadisticas_cplex']]
        
        plt.plot(ejecuciones, llamadas, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Número de Ejecución', fontsize=12, fontweight='bold')
    plt.ylabel('Total de Llamadas a CPLEX', fontsize=12, fontweight='bold')
    plt.title('Llamadas Totales a CPLEX por Ejecución\n(Comparación entre Grupos)', 
              fontsize=14, fontweight='bold')
    plt.legend(loc='best', framealpha=0.9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '02_llamadas_cplex_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 02_llamadas_cplex_comparativo.png")

def grafico_3_tiempo_cplex(todos_datos):
    """Gráfico 3: Tiempo total CPLEX por ejecución"""
    plt.figure(figsize=(12, 8))
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_cplex'] or datos['grupo'] == 0:
            continue  # Grupo 0 no usa CPLEX
        
        ejecuciones = [d['ejecucion'] for d in datos['estadisticas_cplex']]
        tiempos = [d['tiempo_cplex'] for d in datos['estadisticas_cplex']]
        
        plt.plot(ejecuciones, tiempos, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Número de Ejecución', fontsize=12, fontweight='bold')
    plt.ylabel('Tiempo Total CPLEX (segundos)', fontsize=12, fontweight='bold')
    plt.title('Tiempo Total de CPLEX por Ejecución\n(Comparación entre Grupos)', 
              fontsize=14, fontweight='bold')
    plt.legend(loc='best', framealpha=0.9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '03_tiempo_cplex_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 03_tiempo_cplex_comparativo.png")

def grafico_4_evolucion_fitness(todos_datos):
    """Gráfico 4: Evolución del fitness promedio y mejor"""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 12))
    
    # Fitness Promedio (AvgFitness)
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        fitness_prom = [d['avg_fitness'] for d in datos['estadisticas_prom_mej']]
        
        ax1.plot(generaciones, fitness_prom, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    ax1.set_xlabel('Generación', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Fitness Promedio', fontsize=12, fontweight='bold')
    ax1.set_title('Evolución del Fitness Promedio\n(Todos los Grupos)', 
                  fontsize=14, fontweight='bold')
    ax1.legend(loc='best', framealpha=0.9)
    ax1.grid(True, alpha=0.3)
    
    # Fitness Mejor (BestFitness)
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        fitness_mejor = [d['best_fitness'] for d in datos['estadisticas_prom_mej']]
        
        ax2.plot(generaciones, fitness_mejor, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    ax2.set_xlabel('Generación', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Mejor Fitness', fontsize=12, fontweight='bold')
    ax2.set_title('Evolución del Mejor Fitness\n(Todos los Grupos)', 
                  fontsize=14, fontweight='bold')
    ax2.legend(loc='best', framealpha=0.9)
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '04_evolucion_fitness_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 04_evolucion_fitness_comparativo.png")

def grafico_5_evolucion_erp(todos_datos):
    """Gráfico 5: Evolución del ERP promedio y mejor"""
    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 12))
    
    # ERP Promedio (AvgERP)
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        erp_prom = [d['avg_erp'] for d in datos['estadisticas_prom_mej']]
        
        ax1.plot(generaciones, erp_prom, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    ax1.set_xlabel('Generación', fontsize=12, fontweight='bold')
    ax1.set_ylabel('ERP Promedio', fontsize=12, fontweight='bold')
    ax1.set_title('Evolución del ERP Promedio\n(Todos los Grupos)', 
                  fontsize=14, fontweight='bold')
    ax1.legend(loc='best', framealpha=0.9)
    ax1.grid(True, alpha=0.3)
    
    # ERP Mejor (BestERP)
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        erp_mejor = [d['best_erp'] for d in datos['estadisticas_prom_mej']]
        
        ax2.plot(generaciones, erp_mejor, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    ax2.set_xlabel('Generación', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Mejor ERP', fontsize=12, fontweight='bold')
    ax2.set_title('Evolución del Mejor ERP\n(Todos los Grupos)', 
                  fontsize=14, fontweight='bold')
    ax2.legend(loc='best', framealpha=0.9)
    ax2.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '05_evolucion_erp_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 05_evolucion_erp_comparativo.png")

def grafico_6_individuos_evaluados(todos_datos):
    """Gráfico 6: Total de individuos evaluados por generación"""
    plt.figure(figsize=(12, 8))
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        # Asumiendo 15 individuos por generación (población)
        individuos = [gen * 15 for gen in generaciones]
        
        plt.plot(generaciones, individuos, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Generación', fontsize=12, fontweight='bold')
    plt.ylabel('Total de Individuos Evaluados Acumulados', fontsize=12, fontweight='bold')
    plt.title('Individuos Evaluados Acumulados por Generación\n(Todos los Grupos)', 
              fontsize=14, fontweight='bold')
    plt.legend(loc='best', framealpha=0.9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '06_individuos_evaluados_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 06_individuos_evaluados_comparativo.png")

def grafico_7_promedio_llamadas_indiv(todos_datos):
    """Gráfico 7: Promedio de llamadas CPLEX por individuo"""
    plt.figure(figsize=(12, 8))
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_cplex'] or datos['grupo'] == 0:
            continue  # Grupo 0 no usa CPLEX
        
        ejecuciones = [d['ejecucion'] for d in datos['estadisticas_cplex']]
        llamadas_por_indiv = [d['prom_llamadas_indiv'] for d in datos['estadisticas_cplex']]
        
        plt.plot(ejecuciones, llamadas_por_indiv, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Número de Ejecución', fontsize=12, fontweight='bold')
    plt.ylabel('Promedio de Llamadas CPLEX por Individuo', fontsize=12, fontweight='bold')
    plt.title('Promedio de Llamadas CPLEX por Individuo\n(Comparación entre Grupos)', 
              fontsize=14, fontweight='bold')
    plt.legend(loc='best', framealpha=0.9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '07_promedio_llamadas_indiv_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 07_promedio_llamadas_indiv_comparativo.png")

def grafico_8_evolucion_hits(todos_datos):
    """Gráfico 8: Evolución de hits promedio"""
    plt.figure(figsize=(12, 8))
    
    for idx, datos in enumerate(todos_datos):
        if not datos or not datos['estadisticas_prom_mej']:
            continue
        
        generaciones = [d['generacion'] for d in datos['estadisticas_prom_mej']]
        hits = [d['avg_hits'] for d in datos['estadisticas_prom_mej']]
        
        plt.plot(generaciones, hits, marker='o', markersize=3, linewidth=2,
                label=f"Grupo {datos['grupo']}: {datos['presupuesto']}", 
                color=COLORES[idx], alpha=0.8)
    
    plt.xlabel('Generación', fontsize=12, fontweight='bold')
    plt.ylabel('Hits Promedio', fontsize=12, fontweight='bold')
    plt.title('Evolución de Hits Promedio por Generación\n(Todos los Grupos)', 
              fontsize=14, fontweight='bold')
    plt.legend(loc='best', framealpha=0.9)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, '08_evolucion_hits_comparativo.png'), dpi=150, bbox_inches='tight')
    plt.close()
    print("✅ Generado: 08_evolucion_hits_comparativo.png")

def main():
    print("\n" + "="*80)
    print("GENERADOR DE GRÁFICOS COMPARATIVOS SUPERPUESTOS")
    print("="*80 + "\n")
    
    # Cargar datos de todos los grupos
    print("📊 Cargando datos de todos los grupos...")
    todos_datos = []
    for grupo in GRUPOS:
        print(f"   Cargando Grupo {grupo}...", end=" ")
        datos = cargar_datos_grupo(grupo)
        if datos:
            todos_datos.append(datos)
            print("✅")
        else:
            print("❌")
    
    if not todos_datos:
        print("\n❌ ERROR: No se pudieron cargar datos de ningún grupo.")
        return
    
    print(f"\n✅ Se cargaron exitosamente {len(todos_datos)} grupos.\n")
    
    # Generar todos los gráficos
    print("📈 Generando gráficos comparativos...\n")
    
    grafico_1_baseline_tiempos(todos_datos)
    grafico_2_llamadas_cplex(todos_datos)
    grafico_3_tiempo_cplex(todos_datos)
    grafico_4_evolucion_fitness(todos_datos)
    grafico_5_evolucion_erp(todos_datos)
    grafico_6_individuos_evaluados(todos_datos)
    grafico_7_promedio_llamadas_indiv(todos_datos)
    grafico_8_evolucion_hits(todos_datos)
    
    print("\n" + "="*80)
    print(f"✅ COMPLETADO: Se generaron 8 gráficos comparativos en '{OUTPUT_DIR}/'")
    print("="*80 + "\n")

if __name__ == "__main__":
    main()
