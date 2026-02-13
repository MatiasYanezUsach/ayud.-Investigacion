#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Verificación completa de todos los reportes Excel generados
"""

from openpyxl import load_workbook
import os

def verificar_archivo(filename):
    """Verifica un archivo Excel y retorna si tiene datos válidos"""
    if not os.path.exists(filename):
        return False, "Archivo no encontrado"
    
    try:
        wb = load_workbook(filename, data_only=True)
        
        # Buscar hoja de Resumen
        ws = None
        for sheet_name in wb.sheetnames:
            if "Resumen" in sheet_name:
                ws = wb[sheet_name]
                break
        
        if not ws:
            return False, "No se encontró hoja Resumen"
        
        # Buscar métricas clave
        mejor_erp = None
        erp_poblacional = None
        total_individuos = None
        total_ejecuciones = None
        
        for row in range(1, ws.max_row + 1):
            key = ws.cell(row, 1).value
            value = ws.cell(row, 2).value
            
            if key and value is not None:
                key_str = str(key).strip()
                if "Mejor ERP Promedio" in key_str:
                    mejor_erp = float(value)
                elif "ERP Poblacional Promedio" in key_str:
                    erp_poblacional = float(value)
                elif "Total Individuos Evaluados" in key_str:
                    total_individuos = float(value)
                elif "Total Ejecuciones" in key_str:
                    total_ejecuciones = int(value)
        
        # Validar que tenga datos
        if mejor_erp and mejor_erp > 0 and total_ejecuciones and total_ejecuciones > 0:
            return True, f"OK - Mejor ERP: {mejor_erp:.6f}, Ejecuciones: {total_ejecuciones}, Individuos: {int(total_individuos) if total_individuos else 0}"
        else:
            return False, f"Datos incompletos - Mejor ERP: {mejor_erp}, Ejecuciones: {total_ejecuciones}"
        
    except Exception as e:
        return False, f"Error: {str(e)}"

def main():
    print("=" * 80)
    print("VERIFICACIÓN COMPLETA DE REPORTES EXCEL")
    print("=" * 80)
    print()
    
    # Verificar archivos individuales
    print("1. ARCHIVOS INDIVIDUALES POR GRUPO")
    print("-" * 80)
    
    all_ok = True
    for g in range(6):
        filename = f"RESULTADOS_EXPERIMENTO_GRUPO{g}.xlsx"
        ok, msg = verificar_archivo(filename)
        
        status = "✅" if ok else "❌"
        print(f"{status} {filename:40s} {msg}")
        
        if not ok:
            all_ok = False
    
    print()
    
    # Verificar consolidado
    print("2. ARCHIVO CONSOLIDADO")
    print("-" * 80)
    
    filename = "RESULTADOS_EXPERIMENTO_CONSOLIDADO.xlsx"
    if os.path.exists(filename):
        try:
            wb = load_workbook(filename, data_only=True)
            
            # Verificar hoja "Comparación de Grupos"
            if "Comparación de Grupos" in wb.sheetnames:
                ws = wb["Comparación de Grupos"]
                
                # Contar grupos (empezar desde fila 4)
                grupos_count = 0
                row = 4
                while ws.cell(row, 1).value:
                    grupos_count += 1
                    row += 1
                
                if grupos_count == 6:
                    print(f"✅ {filename:40s} OK - {grupos_count} grupos encontrados")
                    
                    # Mostrar datos de cada grupo
                    print()
                    print("   Comparación entre grupos:")
                    print("   " + "-" * 76)
                    for r in range(4, 10):
                        grupo = ws.cell(r, 1).value
                        presupuesto = ws.cell(r, 2).value
                        ejecuciones = ws.cell(r, 3).value
                        mejor_erp = ws.cell(r, 8).value
                        
                        if grupo and mejor_erp:
                            simbolo = "🎯" if r == 4 else "📊"
                            print(f"   {simbolo} {grupo:10s} {presupuesto:20s} Ejecuciones: {ejecuciones:2.0f}  Mejor ERP: {mejor_erp:.6f}")
                else:
                    print(f"❌ {filename:40s} Solo {grupos_count} grupos (esperados: 6)")
                    all_ok = False
            else:
                print(f"❌ {filename:40s} Falta hoja 'Comparación de Grupos'")
                all_ok = False
            
            wb.close()
        except Exception as e:
            print(f"❌ {filename:40s} Error: {str(e)}")
            all_ok = False
    else:
        print(f"❌ {filename:40s} Archivo no encontrado")
        all_ok = False
    
    print()
    print("=" * 80)
    if all_ok:
        print("✅ TODOS LOS ARCHIVOS ESTÁN CORRECTOS Y COMPLETOS")
    else:
        print("❌ ALGUNOS ARCHIVOS TIENEN PROBLEMAS - REVISAR ARRIBA")
    print("=" * 80)

if __name__ == "__main__":
    main()
