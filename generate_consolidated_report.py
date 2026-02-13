"""
Script para generar un reporte Excel consolidado que combine todos los grupos experimentales
desde sus archivos Excel individuales.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import platform

def get_ram_info():
    """Obtiene información de RAM del sistema"""
    try:
        import psutil
        ram_total = psutil.virtual_memory().total
        ram_total_gb = ram_total / (1024 ** 3)
        ram_available_gb = psutil.virtual_memory().available / (1024 ** 3)
        return f"{ram_total_gb:.2f} GB (Disponible: {ram_available_gb:.2f} GB)"
    except ImportError:
        try:
            import subprocess
            if platform.system() == "Windows":
                result = subprocess.run(
                    ['wmic', 'computersystem', 'get', 'TotalPhysicalMemory'],
                    capture_output=True,
                    text=True,
                    check=True
                )
                lines = result.stdout.strip().split('\n')
                if len(lines) > 1:
                    ram_bytes = int(lines[1].strip())
                    ram_gb = ram_bytes / (1024 ** 3)
                    return f"{ram_gb:.2f} GB"
        except:
            pass
        return "N/A"

def read_group_data(group_num):
    """Lee los datos de un grupo desde su Excel"""
    excel_file = f"RESULTADOS_EXPERIMENTO_GRUPO{group_num}.xlsx"
    if not os.path.exists(excel_file):
        return None
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Leer hoja "Resumen General" para obtener información consolidada
        ws_summary = None
        for sheet_name in wb.sheetnames:
            if "Resumen" in sheet_name:
                ws_summary = wb[sheet_name]
                break
        
        summary_data = {}
        if ws_summary:
            for row in range(1, ws_summary.max_row + 1):
                key = ws_summary.cell(row, 1).value
                value = ws_summary.cell(row, 2).value
                if key and value is not None:
                    summary_data[str(key).strip()] = value
        
        # Leer hoja "Estadísticas por Ejecución"
        ws_stats = None
        for sheet_name in wb.sheetnames:
            if "Estad" in sheet_name and "Ejecuci" in sheet_name:
                ws_stats = wb[sheet_name]
                break
        
        data = []
        if ws_stats:
            # Leer datos (saltar encabezados)
            for row in range(3, ws_stats.max_row + 1):
                row_data = {}
                for col in range(1, ws_stats.max_column + 1):
                    header = ws_stats.cell(2, col).value
                    value = ws_stats.cell(row, col).value
                    if header:
                        row_data[header] = value
                if row_data:
                    data.append(row_data)
        
        # Si no hay datos de "Estadísticas por Ejecución", contar ejecuciones manualmente
        if not data and 'Total Ejecuciones:' in summary_data:
            num_ejecuciones = summary_data.get('Total Ejecuciones:', 0)
            # Crear datos sintéticos para el Grupo 0
            for i in range(int(num_ejecuciones)) if isinstance(num_ejecuciones, (int, float)) else range(0):
                data.append({'Ejecución': i})
        
        # Leer hoja "Best Fitness por Ejecución"
        ws_fitness = None
        for sheet_name in wb.sheetnames:
            if "Best Fitness" in sheet_name or "Fitness" in sheet_name:
                ws_fitness = wb[sheet_name]
                break
        
        fitness_data = []
        if ws_fitness:
            for row in range(3, ws_fitness.max_row + 1):
                row_data = {}
                for col in range(1, ws_fitness.max_column + 1):
                    header = ws_fitness.cell(2, col).value
                    value = ws_fitness.cell(row, col).value
                    if header:
                        row_data[header] = value
                if row_data:
                    fitness_data.append(row_data)
        
        # Leer hoja "Estadísticas Promedio y Mejor"
        ws_avg = None
        for sheet_name in wb.sheetnames:
            if "Promedio" in sheet_name or "Prom" in sheet_name:
                ws_avg = wb[sheet_name]
                break
        
        avg_data = {}
        if ws_avg:
            # Buscar filas con datos promedio
            for row in range(2, ws_avg.max_row + 1):
                key = ws_avg.cell(row, 1).value
                if key and ("Promedio" in str(key) or "Mejor" in str(key)):
                    avg_data[key] = {}
                    for col in range(2, ws_avg.max_column + 1):
                        header = ws_avg.cell(1, col).value
                        value = ws_avg.cell(row, col).value
                        if header:
                            avg_data[key][header] = value
        
        return {
            'stats': data,
            'fitness': fitness_data,
            'averages': avg_data,
            'summary': summary_data
        }
    except Exception as e:
        print(f"Error leyendo grupo {group_num}: {e}")
        return None

def generate_consolidated_excel():
    """Genera el reporte Excel consolidado"""
    
    excel_file = "RESULTADOS_EXPERIMENTO_CONSOLIDADO.xlsx"
    print(f"\nGenerando reporte Excel consolidado: {excel_file}")
    print("=" * 80)
    
    # Leer datos de todos los grupos
    grupos = [0, 1, 2, 3, 4, 5]
    grupos_nombres = {
        0: "Grupo 0 - 0% (Sin CPLEX)",
        1: "Grupo 1 - 10%",
        2: "Grupo 2 - 25%",
        3: "Grupo 3 - 50%",
        4: "Grupo 4 - 75%",
        5: "Grupo 5 - 100%"
    }
    
    print("Leyendo datos de todos los grupos...")
    grupos_data = {}
    for g in grupos:
        print(f"  Leyendo Grupo {g}...")
        data = read_group_data(g)
        if data:
            grupos_data[g] = data
            print(f"    [OK] {len(data['stats'])} ejecuciones encontradas")
        else:
            print(f"    [ERROR] No se pudo leer Grupo {g}")
    
    if not grupos_data:
        print("ERROR: No se encontraron datos de ningún grupo")
        return False
    
    # Crear workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ===== HOJA 1: Configuración =====
    print("Creando hoja: Configuración...")
    ws_config = wb.create_sheet("Configuración", 0)
    
    row = 1
    ws_config.merge_cells(f'A{row}:C{row}')
    cell = ws_config.cell(row, 1)
    cell.value = "CONFIGURACIÓN DEL EXPERIMENTO - TODOS LOS GRUPOS"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    ws_config.cell(row, 1).value = "INFORMACIÓN DE LA MÁQUINA"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    ws_config.cell(row, 1).value = "Sistema Operativo:"
    ws_config.cell(row, 2).value = platform.system() + " " + platform.release()
    row += 1
    ws_config.cell(row, 1).value = "Arquitectura:"
    ws_config.cell(row, 2).value = platform.machine()
    row += 1
    ws_config.cell(row, 1).value = "Procesador:"
    ws_config.cell(row, 2).value = platform.processor() if platform.processor() else "N/A"
    row += 1
    ws_config.cell(row, 1).value = "Memoria RAM:"
    ws_config.cell(row, 2).value = get_ram_info()
    row += 2
    
    ws_config.cell(row, 1).value = "GRUPOS EXPERIMENTALES"
    ws_config.cell(row, 1).font = Font(bold=True)
    row += 1
    for g in grupos:
        ws_config.cell(row, 1).value = grupos_nombres[g]
        ws_config.cell(row, 2).value = f"{len(grupos_data.get(g, {}).get('stats', []))} ejecuciones" if g in grupos_data else "No disponible"
        row += 1
    
    ws_config.column_dimensions['A'].width = 35
    ws_config.column_dimensions['B'].width = 40
    
    # ===== HOJA 2: Comparación de Grupos =====
    print("Creando hoja: Comparación de Grupos...")
    ws_compare = wb.create_sheet("Comparación de Grupos")
    
    row = 1
    ws_compare.merge_cells(f'A{row}:H{row}')
    cell = ws_compare.cell(row, 1)
    cell.value = "COMPARACIÓN ENTRE GRUPOS EXPERIMENTALES"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    headers = ["Grupo", "Presupuesto", "Ejecuciones", "Individuos Evaluados (Promedio)", 
               "Llamadas CPLEX (Promedio)", "Tiempo CPLEX Total (Promedio, s)",
               "Fitness ERP (Promedio)", "Fitness ERP (Mejor)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_compare.cell(row, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    for g in grupos:
        if g not in grupos_data:
            continue
        
        data = grupos_data[g]
        stats = data['stats']
        summary_data = data.get('summary', {})
        
        # Para Grupo 0 (sin CPLEX), usar datos del resumen
        if g == 0:
            num_ejecuciones = int(summary_data.get('Total Ejecuciones:', 0))
            avg_indiv = float(summary_data.get('Total Individuos Evaluados:', 0))
            avg_calls = 0
            avg_time = 0
            avg_fitness = float(summary_data.get('ERP Poblacional Promedio:', 0))
            best_fitness = float(summary_data.get('Mejor ERP Promedio:', 0))
        else:
            if not stats:
                continue
            
            # Calcular promedios
            total_indiv = sum(float(s.get('Individuos Evaluados', 0) or 0) for s in stats)
            total_calls = sum(float(s.get('Llamadas CPLEX (Total)', 0) or 0) for s in stats)
            total_time = sum(float(s.get('Tiempo Total CPLEX (s)', 0) or 0) for s in stats)
            
            avg_indiv = total_indiv / len(stats) if stats else 0
            avg_calls = total_calls / len(stats) if stats else 0
            avg_time = total_time / len(stats) if stats else 0
            num_ejecuciones = len(stats)
            
            # Fitness promedio y mejor
            fitness_data = data.get('fitness', [])
            if fitness_data:
                # Obtener última generación de cada ejecución
                last_gen_fitness = []
                for exec_num in range(len(stats)):
                    exec_fitness = [f for f in fitness_data if f.get('Ejecución') == exec_num + 1]
                    if exec_fitness:
                        # Última generación
                        last_gen = max(exec_fitness, key=lambda x: x.get('Generación', 0))
                        erp = last_gen.get('Fitness Estandarizado (ERP)')
                        if erp is not None:
                            try:
                                last_gen_fitness.append(float(erp))
                            except:
                                pass
                
                avg_fitness = sum(last_gen_fitness) / len(last_gen_fitness) if last_gen_fitness else 0
                best_fitness = min(last_gen_fitness) if last_gen_fitness else 0
            else:
                avg_fitness = 0
                best_fitness = 0
        
        # Escribir fila
        ws_compare.cell(row, 1).value = f"Grupo {g}"
        ws_compare.cell(row, 2).value = grupos_nombres[g].split(" - ")[1]
        ws_compare.cell(row, 3).value = num_ejecuciones
        ws_compare.cell(row, 4).value = round(avg_indiv, 2)
        ws_compare.cell(row, 5).value = round(avg_calls, 2)
        ws_compare.cell(row, 6).value = round(avg_time, 6)
        ws_compare.cell(row, 7).value = round(avg_fitness, 6)
        ws_compare.cell(row, 8).value = round(best_fitness, 6)
        
        # Aplicar formato
        for col in range(1, len(headers) + 1):
            cell = ws_compare.cell(row, col)
            cell.border = border
            if col > 3:
                cell.number_format = '0.00' if col < 7 else '0.000000'
        
        row += 1
    
    # Ajustar anchos
    for col in range(1, len(headers) + 1):
        ws_compare.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== HOJA 3: Estadísticas Detalladas por Grupo =====
    print("Creando hoja: Estadísticas Detalladas...")
    ws_detailed = wb.create_sheet("Estadísticas Detalladas")
    
    row = 1
    ws_detailed.merge_cells(f'A{row}:J{row}')
    cell = ws_detailed.cell(row, 1)
    cell.value = "ESTADÍSTICAS DETALLADAS POR GRUPO Y EJECUCIÓN"
    cell.font = title_font
    cell.alignment = Alignment(horizontal='center')
    
    row += 2
    headers = ["Grupo", "Ejecución", "Individuos Evaluados", "Llamadas CPLEX (Total)",
               "Tiempo Total CPLEX (s)", "Promedio Llamadas/Indiv", "Promedio Tiempo/Indiv (s)",
               "Fitness ERP Final", "Mejor Fitness ERP"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_detailed.cell(row, col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    for g in grupos:
        if g not in grupos_data:
            continue
        
        data = grupos_data[g]
        stats = data['stats']
        fitness_data = data.get('fitness', [])
        
        for exec_num, stat in enumerate(stats):
            # Obtener fitness de esta ejecución
            exec_fitness = [f for f in fitness_data if f.get('Ejecución') == exec_num + 1]
            final_fitness = 0
            best_fitness = 0
            if exec_fitness:
                last_gen = max(exec_fitness, key=lambda x: x.get('Generación', 0))
                final_fitness = float(last_gen.get('Fitness Estandarizado (ERP)', 0) or 0)
                best_fitness = min(float(f.get('Fitness Estandarizado (ERP)', 0) or 0) for f in exec_fitness)
            
            ws_detailed.cell(row, 1).value = f"Grupo {g}"
            ws_detailed.cell(row, 2).value = exec_num + 1
            ws_detailed.cell(row, 3).value = float(stat.get('Individuos Evaluados', 0) or 0)
            ws_detailed.cell(row, 4).value = float(stat.get('Llamadas CPLEX (Total)', 0) or 0)
            ws_detailed.cell(row, 5).value = float(stat.get('Tiempo Total CPLEX (s)', 0) or 0)
            ws_detailed.cell(row, 6).value = float(stat.get('Promedio Llamadas/Indiv', 0) or 0)
            ws_detailed.cell(row, 7).value = float(stat.get('Promedio Tiempo/Indiv (s)', 0) or 0)
            ws_detailed.cell(row, 8).value = final_fitness
            ws_detailed.cell(row, 9).value = best_fitness
            
            # Aplicar formato
            for col in range(1, len(headers) + 1):
                cell = ws_detailed.cell(row, col)
                cell.border = border
                if col > 2:
                    cell.number_format = '0.00' if col < 8 else '0.000000'
            
            row += 1
    
    # Ajustar anchos
    for col in range(1, len(headers) + 1):
        ws_detailed.column_dimensions[get_column_letter(col)].width = 18
    
    # Guardar
    print(f"\nGuardando archivo: {excel_file}")
    wb.save(excel_file)
    print(f"[OK] Archivo Excel consolidado creado: {excel_file}")
    
    return True

if __name__ == "__main__":
    generate_consolidated_excel()

