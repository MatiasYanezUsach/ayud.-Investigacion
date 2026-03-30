"""
Curvas de convergencia para prueba de tamaño de población.
Lee resultados de out/prueba_poblacion/ y genera gráficos comparativos.

Directorios esperados:
  out/prueba_poblacion/popXXX_cplexYY/evolution0/job.0.BestFitness.csv
"""
import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.rcParams['font.family'] = 'DejaVu Sans'

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PRUEBA_DIR = os.path.join(BASE_DIR, "out", "prueba_poblacion")
OUT_DIR    = os.path.join(BASE_DIR, "out")

POBLACIONES = [100, 75, 50, 25, 20, 15, 10, 5, 1]
COLORES_POP = {
    100: "#e41a1c",
    75:  "#ff7f00",
    50:  "#4daf4a",
    25:  "#377eb8",
    20:  "#984ea3",
    15:  "#a65628",
    10:  "#f781bf",
    5:   "#999999",
    1:   "#000000",
}


def leer_fitness(outdir: str):
    """Lee fitness por generación desde job.0.BestFitness.csv o Statistics.out."""
    evo_dir = os.path.join(outdir, "evolution0")

    # Opción 1: archivo CSV con prefijo job
    csv_path = os.path.join(evo_dir, "job.0.BestFitness.csv")
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, sep=";", decimal=",")
            df.columns = [c.strip() for c in df.columns]
            return df["Standarized"].values
        except Exception as e:
            print(f"  Error leyendo {csv_path}: {e}")

    # Opción 2: Statistics.out (sin prefijo job)
    stat_path = os.path.join(evo_dir, "Statistics.out")
    if os.path.exists(stat_path):
        try:
            fitness_vals = []
            capture_next = False
            with open(stat_path, "r", encoding="utf-8", errors="replace") as f:
                for line in f:
                    if line.startswith("Best Individual:"):
                        capture_next = True
                    elif capture_next and "Fitness: Standardized=" in line:
                        part = line.split("Standardized=")[1].split()[0]
                        fitness_vals.append(float(part))
                        capture_next = False
            if fitness_vals:
                return np.array(fitness_vals)
        except Exception as e:
            print(f"  Error leyendo {stat_path}: {e}")

    print(f"  Sin datos en: {evo_dir}")
    return None


def graficar(cplex_pct: int):
    """Genera gráfico comparativo de poblaciones para un % de CPLEX dado."""
    fig, ax = plt.subplots(figsize=(11, 5))
    encontrados = 0

    for pop in POBLACIONES:
        outdir = os.path.join(PRUEBA_DIR, f"pop{pop}_cplex{cplex_pct}")
        fitness = leer_fitness(outdir)
        if fitness is None:
            continue

        gen = np.arange(len(fitness))
        ax.plot(gen, fitness, color=COLORES_POP[pop], linewidth=2,
                label=f"Población {pop}")

        # Marcar generación de convergencia
        mejoras = np.abs(np.diff(fitness))
        if np.any(mejoras > 0.001):
            conv_gen = np.where(mejoras > 0.001)[0][-1] + 1
            ax.axvline(x=conv_gen, color=COLORES_POP[pop],
                       linestyle="--", alpha=0.4)
        encontrados += 1

    if encontrados == 0:
        print(f"  Sin datos para CPLEX={cplex_pct}%. Ejecuta run_prueba_poblacion.bat primero.")
        plt.close()
        return None

    titulo = f"Convergencia por Tamaño de Población — CPLEX {cplex_pct}%"
    ax.set_title(titulo, fontsize=13)
    ax.set_xlabel("Generación", fontsize=11)
    ax.set_ylabel("Fitness Estandarizado (menor = mejor)", fontsize=11)
    ax.legend(fontsize=10)
    ax.grid(True, alpha=0.3)
    ax.invert_yaxis()

    outfile = os.path.join(OUT_DIR, f"convergencia_poblacion_cplex{cplex_pct}.png")
    plt.tight_layout()
    plt.savefig(outfile, dpi=150)
    plt.show()
    print(f"  Guardado: {outfile}")
    return outfile


def graficar_comparativo_4():
    """Genera figura con 2 subplots: CPLEX 0% y CPLEX 100% lado a lado."""
    fig, axes = plt.subplots(1, 2, figsize=(16, 5), sharey=True)

    for ax, cplex_pct in zip(axes, [0, 100]):
        encontrados = 0
        for pop in POBLACIONES:
            outdir = os.path.join(PRUEBA_DIR, f"pop{pop}_cplex{cplex_pct}")
            fitness = leer_fitness(outdir)
            if fitness is None:
                continue
            gen = np.arange(len(fitness))
            ax.plot(gen, fitness, color=COLORES_POP[pop], linewidth=2,
                    label=f"Población {pop}")
            encontrados += 1

        if encontrados == 0:
            ax.text(0.5, 0.5, "Sin datos", ha="center", va="center",
                    transform=ax.transAxes, fontsize=12)
        ax.set_title(f"CPLEX {cplex_pct}%", fontsize=12)
        ax.set_xlabel("Generación", fontsize=10)
        ax.grid(True, alpha=0.3)
        ax.invert_yaxis()
        ax.legend(fontsize=9)

    axes[0].set_ylabel("Fitness Estandarizado (menor = mejor)", fontsize=10)
    fig.suptitle("Convergencia por Tamaño de Población", fontsize=14, fontweight="bold")

    outfile = os.path.join(OUT_DIR, "convergencia_poblacion_comparativo.png")
    plt.tight_layout()
    plt.savefig(outfile, dpi=150)
    plt.show()
    print(f"Guardado: {outfile}")


def generar_excel():
    """Genera Excel con fitness por generación para todas las combinaciones."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  openpyxl no instalado. Ejecuta: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    resumen_data = []

    for cplex_pct in [0, 100]:
        ws = wb.create_sheet(title=f"CPLEX {cplex_pct}%")

        # Encabezado
        ws.cell(1, 1, "Población").font = Font(bold=True)
        ws.cell(1, 2, "Gen. Convergencia").font = Font(bold=True)
        ws.cell(1, 3, "Fitness Inicial").font = Font(bold=True)
        ws.cell(1, 4, "Fitness Final").font = Font(bold=True)
        ws.cell(1, 5, "Mejora Total").font = Font(bold=True)

        row = 2
        for pop in POBLACIONES:
            outdir = os.path.join(PRUEBA_DIR, f"pop{pop}_cplex{cplex_pct}")
            fitness = leer_fitness(outdir)
            if fitness is None:
                continue

            mejoras = np.abs(np.diff(fitness))
            conv_gen = int(np.where(mejoras > 0.001)[0][-1]) + 1 if np.any(mejoras > 0.001) else 0
            fit_ini  = float(fitness[0])
            fit_fin  = float(fitness[-1])
            mejora   = fit_ini - fit_fin

            ws.cell(row, 1, pop)
            ws.cell(row, 2, conv_gen)
            ws.cell(row, 3, round(fit_ini, 6))
            ws.cell(row, 4, round(fit_fin, 6))
            ws.cell(row, 5, round(mejora, 6))
            row += 1

            resumen_data.append({
                "CPLEX": f"{cplex_pct}%",
                "Población": pop,
                "Gen. Convergencia": conv_gen,
                "Fitness Inicial": round(fit_ini, 6),
                "Fitness Final": round(fit_fin, 6),
                "Mejora Total": round(mejora, 6),
            })

        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20

    # Hoja resumen con todos los datos
    ws_res = wb.create_sheet(title="Resumen", index=0)
    headers = ["CPLEX", "Población", "Gen. Convergencia", "Fitness Inicial", "Fitness Final", "Mejora Total"]
    for c, h in enumerate(headers, 1):
        cell = ws_res.cell(1, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.font = Font(bold=True, color="FFFFFF")
        ws_res.column_dimensions[get_column_letter(c)].width = 20

    for r, row_data in enumerate(resumen_data, 2):
        for c, key in enumerate(headers, 1):
            ws_res.cell(r, c, row_data[key])

    outfile = os.path.join(BASE_DIR, "reporte_prueba_poblacion.xlsx")
    wb.save(outfile)
    print(f"  Excel guardado: {outfile}")


if __name__ == "__main__":
    print("=== Curvas de convergencia — Prueba de población ===\n")

    print("Generando gráfico CPLEX 0%...")
    graficar(0)

    print("\nGenerando gráfico CPLEX 100%...")
    graficar(100)

    print("\nGenerando gráfico comparativo (0% vs 100%)...")
    graficar_comparativo_4()

    print("\nGenerando Excel...")
    generar_excel()

    print("\nListo.")
