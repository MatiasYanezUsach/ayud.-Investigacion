#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Verificar que el reporte consolidado incluye datos del Grupo 0
"""

from openpyxl import load_workbook

def verificar_consolidado():
    wb = load_workbook("RESULTADOS_EXPERIMENTO_CONSOLIDADO.xlsx")
    
    # Verificar hoja "Comparación de Grupos"
    if "Comparación de Grupos" in wb.sheetnames:
        ws = wb["Comparación de Grupos"]
        print("=" * 80)
        print("HOJA: Comparación de Grupos")
        print("=" * 80)
        
        # Leer encabezados (fila 3)
        headers = []
        for col in range(1, 9):
            val = ws.cell(3, col).value
            if val:
                headers.append(str(val))
        
        print("\nEncabezados:")
        print(f"  {' | '.join(headers)}")
        print()
        
        # Leer datos de todos los grupos (empezando desde fila 4)
        print("Datos de los grupos:\n")
        row = 4
        while ws.cell(row, 1).value:
            grupo = ws.cell(row, 1).value
            presupuesto = ws.cell(row, 2).value
            ejecuciones = ws.cell(row, 3).value
            individuos = ws.cell(row, 4).value
            llamadas = ws.cell(row, 5).value
            tiempo = ws.cell(row, 6).value
            avg_fitness = ws.cell(row, 7).value
            best_fitness = ws.cell(row, 8).value
            
            print(f"{grupo}:")
            print(f"  Presupuesto CPLEX: {presupuesto}")
            print(f"  Ejecuciones: {ejecuciones}")
            print(f"  Individuos evaluados (prom): {individuos}")
            print(f"  Llamadas CPLEX (prom): {llamadas}")
            print(f"  Tiempo CPLEX (prom, s): {tiempo}")
            print(f"  ERP Promedio: {avg_fitness}")
            print(f"  ERP Mejor: {best_fitness}")
            print()
            
            row += 1
        
        # Análisis de resultados
        print("=" * 80)
        print("ANÁLISIS: Comparación con Grupo 0 (BASELINE)")
        print("=" * 80)
        
        # Obtener mejor ERP del Grupo 0
        grupo0_best = ws.cell(4, 8).value  # Grupo 0 está en fila 4, columna 8
        if grupo0_best:
            print(f"\nGrupo 0 (0% CPLEX) - Mejor ERP: {grupo0_best:.6f} ← BASELINE\n")
            
            # Comparar con otros grupos
            for r in range(5, 10):  # Grupos 1-5
                if ws.cell(r, 1).value:
                    grupo_name = ws.cell(r, 1).value
                    grupo_best = ws.cell(r, 8).value
                    if grupo_best:
                        mejora = ((grupo0_best - grupo_best) / grupo0_best) * 100
                        simbolo = "✓" if mejora > 0 else "✗"
                        direccion = "MEJOR" if mejora > 0 else "PEOR"
                        print(f"{simbolo} {grupo_name}: {grupo_best:.6f} ({mejora:+.2f}% vs Grupo 0) - {direccion}")
    
    else:
        print("ERROR: No se encontró la hoja 'Comparación de Grupos'")
    
    wb.close()

if __name__ == "__main__":
    verificar_consolidado()
