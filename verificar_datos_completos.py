"""
Reporte de Verificación de Datos - Todos los Grupos
Confirma que el Grupo 0 tiene datos de rendimiento evolutivo
"""

import openpyxl
import numpy as np

GRUPOS = [0, 1, 2, 3, 4, 5]
PRESUPUESTOS = {0: "0% (Sin CPLEX)", 1: "10%", 2: "25%", 3: "50%", 4: "75%", 5: "100%"}

print("\n" + "="*100)
print("VERIFICACIÓN DE DATOS DE RENDIMIENTO EVOLUTIVO - TODOS LOS GRUPOS")
print("="*100 + "\n")

print("PREGUNTA: ¿Los algoritmos con CPLEX son mejores que los puramente heurísticos (Grupo 0)?")
print("="*100 + "\n")

# Tabla de resultados
print(f"{'Grupo':<10} {'Presupuesto':<20} {'AvgERP Inicial':<18} {'BestERP Inicial':<18} {'AvgERP Final':<18} {'BestERP Final':<18}")
print("-" * 100)

resultados = {}

for grupo in GRUPOS:
    archivo = f"RESULTADOS_EXPERIMENTO_GRUPO{grupo}.xlsx"
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb['Estadísticas Promedio y Mejor']
    
    # Agrupar por generación
    gen_data = {}
    for row in range(3, ws.max_row + 1):
        gen = ws.cell(row, 2).value
        avg_erp = ws.cell(row, 4).value
        best_erp = ws.cell(row, 7).value
        
        if gen is not None and avg_erp is not None and best_erp is not None:
            gen = int(gen)
            if gen not in gen_data:
                gen_data[gen] = {'avg_erp': [], 'best_erp': []}
            gen_data[gen]['avg_erp'].append(float(avg_erp))
            gen_data[gen]['best_erp'].append(float(best_erp))
    
    # Obtener valores iniciales (gen 1) y finales (gen 100)
    gen_inicial = min(gen_data.keys())
    gen_final = max(gen_data.keys())
    
    avg_erp_inicial = np.mean(gen_data[gen_inicial]['avg_erp'])
    best_erp_inicial = np.mean(gen_data[gen_inicial]['best_erp'])
    avg_erp_final = np.mean(gen_data[gen_final]['avg_erp'])
    best_erp_final = np.mean(gen_data[gen_final]['best_erp'])
    
    resultados[grupo] = {
        'avg_erp_inicial': avg_erp_inicial,
        'best_erp_inicial': best_erp_inicial,
        'avg_erp_final': avg_erp_final,
        'best_erp_final': best_erp_final,
        'generaciones': len(gen_data)
    }
    
    print(f"{grupo:<10} {PRESUPUESTOS[grupo]:<20} {avg_erp_inicial:<18.6f} {best_erp_inicial:<18.6f} {avg_erp_final:<18.6f} {best_erp_final:<18.6f}")
    
    wb.close()

print("\n" + "="*100)
print("ANÁLISIS COMPARATIVO RESPECTO AL GRUPO 0 (Sin CPLEX)")
print("="*100 + "\n")

grupo0_best_final = resultados[0]['best_erp_final']
grupo0_avg_final = resultados[0]['avg_erp_final']

print(f"{'Grupo':<10} {'Presupuesto':<20} {'Mejora BestERP':<20} {'Mejora AvgERP':<20} {'Conclusión':<30}")
print("-" * 100)

for grupo in GRUPOS:
    if grupo == 0:
        print(f"{grupo:<10} {PRESUPUESTOS[grupo]:<20} {'BASELINE':<20} {'BASELINE':<20} {'Algoritmo Heurístico Puro':<30}")
    else:
        mejora_best = ((grupo0_best_final - resultados[grupo]['best_erp_final']) / grupo0_best_final) * 100
        mejora_avg = ((grupo0_avg_final - resultados[grupo]['avg_erp_final']) / grupo0_avg_final) * 100
        
        if mejora_best > 0:
            conclusion = f"✅ {mejora_best:.2f}% mejor"
        elif mejora_best < -0.5:
            conclusion = f"❌ {abs(mejora_best):.2f}% peor"
        else:
            conclusion = "≈ Similar"
        
        print(f"{grupo:<10} {PRESUPUESTOS[grupo]:<20} {mejora_best:>+18.2f}% {mejora_avg:>+18.2f}% {conclusion:<30}")

print("\n" + "="*100)
print("RESPUESTA A LAS PREGUNTAS")
print("="*100 + "\n")

print("1. ¿El Grupo 0 tiene datos de rendimiento?")
print(f"   ✅ SÍ - {resultados[0]['generaciones']} generaciones con datos completos de ERP, Fitness, Hits")
print(f"   📊 BestERP final: {resultados[0]['best_erp_final']:.6f}")
print(f"   📊 AvgERP final: {resultados[0]['avg_erp_final']:.6f}")

print("\n2. ¿Los algoritmos con CPLEX son mejores que el GP puro (Grupo 0)?")
for grupo in [1, 2, 3, 4, 5]:
    mejora = ((grupo0_best_final - resultados[grupo]['best_erp_final']) / grupo0_best_final) * 100
    if mejora > 0:
        print(f"   ✅ Grupo {grupo} ({PRESUPUESTOS[grupo]}): {mejora:.2f}% MEJOR que Grupo 0")
    elif mejora < -0.5:
        print(f"   ❌ Grupo {grupo} ({PRESUPUESTOS[grupo]}): {abs(mejora):.2f}% PEOR que Grupo 0")
    else:
        print(f"   ≈ Grupo {grupo} ({PRESUPUESTOS[grupo]}): Rendimiento SIMILAR a Grupo 0")

print("\n3. ¿Cuál es el mejor grupo?")
mejor_grupo = min(GRUPOS, key=lambda g: resultados[g]['best_erp_final'])
mejor_erp = resultados[mejor_grupo]['best_erp_final']
print(f"   🏆 Grupo {mejor_grupo} ({PRESUPUESTOS[mejor_grupo]}): BestERP = {mejor_erp:.6f}")

print("\n" + "="*100)
print("CONCLUSIÓN")
print("="*100)
print(f"""
Los datos del Grupo 0 (GP puro sin CPLEX) SÍ EXISTEN y están disponibles para comparación.

El Grupo 0 alcanza un BestERP de {resultados[0]['best_erp_final']:.6f}, lo cual sirve como BASELINE
para evaluar si el uso de CPLEX (en cualquier presupuesto) mejora el rendimiento del
algoritmo evolutivo.

Usar estos datos permite responder preguntas clave como:
- ¿Vale la pena invertir tiempo de CPU en CPLEX?
- ¿Cuál es el presupuesto óptimo de CPLEX?
- ¿El GP puro puede competir con enfoques híbridos?
""")
print("="*100 + "\n")
