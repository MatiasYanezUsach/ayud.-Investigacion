import openpyxl
import matplotlib.pyplot as plt
import numpy as np

# Cargar datos de todos los grupos
GRUPOS = [0, 1, 2, 3, 4, 5]
PRESUPUESTOS = {0: "0% (Sin CPLEX)", 1: "10%", 2: "25%", 3: "50%", 4: "75%", 5: "100%"}
COLORES = ['#e41a1c', '#377eb8', '#4daf4a', '#984ea3', '#ff7f00', '#f781bf']

# Test: Cargar solo ERP para verificar Grupo 0
erp_datos = {}

for grupo in GRUPOS:
    archivo = f"RESULTADOS_EXPERIMENTO_GRUPO{grupo}.xlsx"
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb['Estadísticas Promedio y Mejor']
    
    # Agrupar por generación
    gen_data = {}
    for row in range(3, ws.max_row + 1):
        gen = ws.cell(row, 2).value
        best_erp = ws.cell(row, 7).value
        
        if gen is not None and best_erp is not None:
            gen = int(gen)
            if gen not in gen_data:
                gen_data[gen] = []
            gen_data[gen].append(float(best_erp))
    
    # Calcular promedio por generación
    erp_datos[grupo] = {
        'generaciones': sorted(gen_data.keys()),
        'best_erp': [np.mean(gen_data[g]) for g in sorted(gen_data.keys())]
    }
    
    print(f"Grupo {grupo}: {len(gen_data)} generaciones, BestERP final = {np.mean(gen_data[max(gen_data.keys())]):.4f}")
    wb.close()

# Crear gráfico de prueba
plt.figure(figsize=(12, 8))

for idx, grupo in enumerate(GRUPOS):
    datos = erp_datos[grupo]
    plt.plot(datos['generaciones'], datos['best_erp'], 
             marker='o', markersize=2, linewidth=2,
             label=f"Grupo {grupo}: {PRESUPUESTOS[grupo]}", 
             color=COLORES[idx], alpha=0.8)

plt.xlabel('Generación', fontsize=12, fontweight='bold')
plt.ylabel('Mejor ERP (Error Relativo Promedio)', fontsize=12, fontweight='bold')
plt.title('TEST: Evolución del Mejor ERP\n(Verificando que Grupo 0 aparece)', fontsize=14, fontweight='bold')
plt.legend(loc='best', framealpha=0.9)
plt.grid(True, alpha=0.3)
plt.tight_layout()
plt.savefig('TEST_grupo0_erp.png', dpi=150, bbox_inches='tight')
print(f"\n✅ Gráfico de prueba guardado: TEST_grupo0_erp.png")
plt.close()
