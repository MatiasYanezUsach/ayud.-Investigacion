import openpyxl

wb = openpyxl.load_workbook("RESULTADOS_EXPERIMENTO_GRUPO0.xlsx", data_only=True)
ws = wb["Estadísticas Promedio y Mejor"]
total_filas = ws.max_row - 2
print(f"Total filas de datos Grupo 0: {total_filas}")
print(f"Esperado: 30 ejecuciones × 101 generaciones = 3030 filas")
print(f"Última fila: Ejec={ws.cell(ws.max_row, 1).value}, Gen={ws.cell(ws.max_row, 2).value}")

# Contar generaciones únicas
generaciones = set()
for i in range(3, ws.max_row + 1):
    gen = ws.cell(i, 2).value
    if gen is not None:
        generaciones.add(int(gen))

print(f"\nGeneraciones únicas: {len(generaciones)} (min={min(generaciones)}, max={max(generaciones)})")
print(f"\nPrimeros 5 valores de Generación 0:")
count = 0
for i in range(3, ws.max_row + 1):
    if ws.cell(i, 2).value == 0:
        print(f"  Ejec={ws.cell(i, 1).value}, AvgERP={ws.cell(i, 4).value}, BestERP={ws.cell(i, 7).value}")
        count += 1
        if count >= 5:
            break

print(f"\nÚltimos 5 valores de Generación 100:")
rows_gen100 = []
for i in range(3, ws.max_row + 1):
    if ws.cell(i, 2).value == 100:
        rows_gen100.append(i)

for i in rows_gen100[-5:]:
    print(f"  Ejec={ws.cell(i, 1).value}, AvgERP={ws.cell(i, 4).value}, BestERP={ws.cell(i, 7).value}")

wb.close()
